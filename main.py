"""
main.py — DroneView UTM Research Analytics | FastAPI Backend
============================================================
This is the entire backend in one file.
It loads simulation data (Excel or PostgreSQL), holds it in memory,
and serves JSON to the frontend via REST API endpoints.

Run:
    python -m uvicorn main:app --reload --port 8000
    Then open  http://127.0.0.1:8000

How it is organised (top to bottom):
  1. App setup & CORS
  2. JSON serialiser         numpy-safe; NaN/Inf → null
  3. Parsers                 battery %, speed, layer → altitude metres
  4. Column maps             DRONE_MAP / COLL_MAP
                             Maps Excel column headers to internal field names
  5. load_excel()            Reads "Run 1", "Run 2"… sheets + "Collision Log"
  6. load_postgres()         Reads drone_summary_3d + collision_log_3d tables
  7. enrich()                Derives computed columns after loading
                             bat_u, spd, cx, cy, layer, alt, dp, da, eff, duration
  8. compute_kpis()          Calculates the 7 headline numbers shown at the top
  9. Chart data builders     One function per chart — called by the route handlers
                             Each returns a plain dict that becomes JSON
 10. _get()                  Filters the in-memory DataFrames by trial + path runs
 11. API routes              @app.get / @app.post endpoints

In-memory session (_S dict):
    _S["drones"]      pd.DataFrame — one row per drone flight
    _S["collisions"]  pd.DataFrame — one row per collision event
    _S["trials"]      list[str]    — unique trial IDs found in the data
    All data is loaded once on /api/connect/* and stays in memory.
    Restarting the server requires reconnecting.

Excel workbook format expected:
    Sheet "Run 1", "Run 2", …
        Row 1: title (e.g. "Simulation Run 1")
        Row 2: column headers  (matched via DRONE_MAP)
        Row 3+: data
    Sheet "Collision Log"
        Row 1: column headers  (matched via COLL_MAP)
        Row 2+: data

PostgreSQL tables expected:
    drone_summary_3d   joined with simulation_runs on run_id
    collision_log_3d   joined with simulation_runs on run_id
    simulation_runs    must have a run_label column → used as trial_id
"""
import io, json, re, os
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import warnings
warnings.filterwarnings("ignore")
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles

# ── App ───────────────────────────────────────────────────────────────────────
app = FastAPI(title="UTM Research Analytics", version="1.0")
app.add_middleware(CORSMiddleware,
    allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

from fastapi import Request
from fastapi.responses import Response

@app.middleware("http")
async def no_cache(request: Request, call_next):
    response = await call_next(request)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

FRONTEND = Path(__file__).resolve().parent / "frontend"
if not FRONTEND.exists():
    FRONTEND.mkdir(parents=True)
    (FRONTEND / "css").mkdir()
    (FRONTEND / "js").mkdir()
    (FRONTEND / "assets").mkdir()
app.mount("/static", StaticFiles(directory=str(FRONTEND)), name="static")

# In-memory session (single-user research tool)
_S: dict = {
    "drones": pd.DataFrame(),
    "collisions": pd.DataFrame(),
    "source": "",
    "trials": [],
}

# ── JSON serialiser ───────────────────────────────────────────────────────────
class _Enc(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, (np.integer,)):  return int(o)
        if isinstance(o, (np.floating,)): return None if (np.isnan(o) or np.isinf(o)) else float(o)
        if isinstance(o, np.ndarray):     return o.tolist()
        return super().default(o)

def J(obj):
    return json.loads(json.dumps(obj, cls=_Enc))

# ── Parsers ───────────────────────────────────────────────────────────────────
def parse_battery(v):
    if pd.isna(v): return np.nan
    try: return float(str(v).replace("%", "").replace(",", ".").strip())
    except: return np.nan

def parse_speed(v):
    if pd.isna(v): return np.nan
    try: return float(str(v).replace("×", "").replace("x", "").strip())
    except: return np.nan

def layer_to_alt(l):
    try: return (int(float(l)) - 1) * 50.0
    except: return np.nan

# ── Column maps ───────────────────────────────────────────────────────────────
DRONE_MAP = {
    "Drone": "drone_id", "Start Time": "start_time", "End Time": "end_time",
    "Flight Status": "flight_status", "Collision Severity": "collision_severity",
    "Vehicle": "vehicle", "Path Label": "path_label", "Drone Speed": "drone_speed",
    "Drone Source": "src", "Drone Destination": "dst",
    "Drone Coord X": "coord_x", "Drone Coord Y": "coord_y",
    "Drone Layer": "drone_layer", "Drone Altitude (m)": "drone_altitude_m",
    "Crashed With Drone": "crashed_with_drone",
    "Battery at Start": "battery_start", "Battery at End": "battery_end",
    "Battery Consumed": "battery_consumed",
    "Distance Planned": "distance_planned", "Distance Actual": "distance_actual",
}
COLL_MAP = {
    "Event ID": "event_id", "Path Run": "path_run", "Path Label": "path_label",
    "Timestamp": "timestamp", "Grid Tick": "grid_tick", "Type": "type",
    "Severity": "severity", "Drone A ID": "drone_a_id", "Drone B ID": "drone_b_id",
    "Drone A Vehicle": "drone_a_vehicle", "Drone B Vehicle": "drone_b_vehicle",
    "Drone A Speed": "drone_a_speed", "Drone B Speed": "drone_b_speed",
    "Drone A Coord X": "drone_a_coord_x", "Drone A Coord Y": "drone_a_coord_y",
    "Drone B Coord X": "drone_b_coord_x", "Drone B Coord Y": "drone_b_coord_y",
    "Drone A Layer": "drone_a_layer", "Drone B Layer": "drone_b_layer",
    "Drone A Altitude (m)": "drone_a_altitude_m", "Drone B Altitude (m)": "drone_b_altitude_m",
    "Layer Diff": "layer_diff",
    "Drone A Battery Start": "drone_a_battery_start", "Drone B Battery Start": "drone_b_battery_start",
    "Drone A Battery End": "drone_a_battery_end", "Drone B Battery End": "drone_b_battery_end",
    "Drone A Battery Consumed": "drone_a_battery_consumed", "Drone B Battery Consumed": "drone_b_battery_consumed",
    "Drone A Distance Actual": "drone_a_distance_actual", "Drone B Distance Actual": "drone_b_distance_actual",
}

# ── Excel loader ──────────────────────────────────────────────────────────────
def load_excel(path_or_file) -> dict:
    import openpyxl
    wb     = openpyxl.load_workbook(path_or_file, data_only=True)
    sheets = wb.sheetnames
    fname  = getattr(path_or_file, "filename",
               getattr(path_or_file, "name", "excel_import"))

    run_sheets  = [s for s in sheets if s.startswith("Run ")]
    drone_parts = []
    for sname in run_sheets:
        ws   = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3: continue
        title = str(rows[0][0]) if rows[0][0] else sname
        m     = re.search(r"Run\s+(\d+)", title)
        prn   = int(m.group(1)) if m else 1
        hdrs  = [str(c).strip() if c else "" for c in rows[1]]
        for row in rows[2:]:
            if not row[0] or str(row[0]).strip() == "": continue
            rec = {DRONE_MAP.get(k, k): v for k, v in zip(hdrs, row)}
            rec["path_run"] = prn
            rec["trial_id"] = str(fname)
            drone_parts.append(rec)

    df_d = pd.DataFrame(drone_parts) if drone_parts else pd.DataFrame()

    coll_parts = []
    if "Collision Log" in sheets:
        ws   = wb["Collision Log"]
        rows = list(ws.iter_rows(values_only=True))
        hdrs = [str(c).strip() if c else "" for c in rows[0]]
        for row in rows[1:]:
            if not row[0]: continue
            rec = {COLL_MAP.get(k, k): v for k, v in zip(hdrs, row)}
            rec["trial_id"] = str(fname)
            coll_parts.append(rec)

    df_c = pd.DataFrame(coll_parts) if coll_parts else pd.DataFrame()

    for df in [df_d, df_c]:
        if not df.empty and df.columns.duplicated().any():
            cols = pd.Index([c for i, c in enumerate(df.columns)
                             if c not in df.columns[:i]])
            df = df.loc[:, ~df.columns.duplicated(keep="first")]

    return {"drones": df_d, "collisions": df_c, "source": str(fname)}


# ── PostgreSQL loader ─────────────────────────────────────────────────────────
def load_postgres(host, port, dbname, user, password) -> dict:
    from sqlalchemy import create_engine
    eng = create_engine(
        f"postgresql+psycopg2://{user}:{password}@{host}:{port}/{dbname}",
        connect_args={"connect_timeout": 10})
    df_d = pd.read_sql(
        "SELECT d.*, r.run_label AS trial_id "
        "FROM drone_summary_3d d JOIN simulation_runs r USING (run_id)", eng)
    df_c = pd.read_sql(
        "SELECT c.*, r.run_label AS trial_id "
        "FROM collision_log_3d c JOIN simulation_runs r USING (run_id)", eng)
    for df in [df_d, df_c]:
        if df.columns.duplicated().any():
            df = df.loc[:, ~df.columns.duplicated(keep="first")]
    return {"drones": df_d, "collisions": df_c, "source": "postgresql"}


# ── Enrich ────────────────────────────────────────────────────────────────────
def enrich(df_d: pd.DataFrame, df_c: pd.DataFrame):
    d = df_d.copy()
    c = df_c.copy()
    if not d.empty:
        d["bat_s"]    = d["battery_start"].apply(parse_battery)
        d["bat_e"]    = d["battery_end"].apply(parse_battery)
        d["bat_u"]    = d["battery_consumed"].apply(parse_battery)
        d["spd"]      = d["drone_speed"].apply(parse_speed)
        d["cx"]       = pd.to_numeric(d["coord_x"],           errors="coerce")
        d["cy"]       = pd.to_numeric(d["coord_y"],           errors="coerce")
        d["layer"]    = pd.to_numeric(d["drone_layer"],       errors="coerce")
        d["alt"]      = d["layer"].apply(layer_to_alt)
        d["dp"]       = pd.to_numeric(d["distance_planned"],  errors="coerce")
        d["da"]       = pd.to_numeric(d["distance_actual"],   errors="coerce")
        d["eff"]      = (d["dp"] / d["da"].replace(0, np.nan)).round(3)
        d["start_dt"] = pd.to_datetime(d["start_time"], errors="coerce")
        d["end_dt"]   = pd.to_datetime(d["end_time"],   errors="coerce")
        d["duration"] = (d["end_dt"] - d["start_dt"]).dt.total_seconds()
    if not c.empty:
        c["gtn"] = pd.to_numeric(c["grid_tick"], errors="coerce")
        for col in ["drone_a_coord_x", "drone_a_coord_y",
                    "drone_b_coord_x", "drone_b_coord_y",
                    "drone_a_layer",   "drone_b_layer"]:
            c[col] = pd.to_numeric(c[col], errors="coerce")
        c["bat_a"] = c["drone_a_battery_start"].apply(parse_battery)
        c["bat_b"] = c["drone_b_battery_start"].apply(parse_battery)
        c["da_a"]  = pd.to_numeric(c["drone_a_distance_actual"], errors="coerce")
        c["da_b"]  = pd.to_numeric(c["drone_b_distance_actual"], errors="coerce")
    return d, c


# ── KPIs ──────────────────────────────────────────────────────────────────────
def compute_kpis(d, c):
    N = len(d)
    if N == 0:
        return dict(N=0, n_ok=0, n_crash=0, n_batt=0, n_canc=0,
                    n_coll=0, n_dir=0, n_prox=0, n_nm=0,
                    comp_pct=0.0, crash_pct=0.0,
                    collision_free_pct=100.0, avg_bat_used=0.0,
                    avg_efficiency=0.0)
    n_ok    = int(d["flight_status"].str.startswith("Complete").sum())
    n_crash = int(d["flight_status"].str.startswith("Collision").sum())
    n_batt  = int(d["flight_status"].str.startswith("Incomplete").sum())
    n_canc  = int(d["flight_status"].str.startswith("Cancelled").sum())
    n_coll  = len(c)
    n_dir   = int((c["type"] == "Direct").sum())    if n_coll else 0
    n_prox  = int((c["type"] == "Proximity").sum()) if n_coll else 0
    n_nm    = int((c["type"] == "Near Miss").sum())  if n_coll else 0
    return dict(
        N=N, n_ok=n_ok, n_crash=n_crash, n_batt=n_batt, n_canc=n_canc,
        n_coll=n_coll, n_dir=n_dir, n_prox=n_prox, n_nm=n_nm,
        comp_pct=round(n_ok / N * 100, 1),
        crash_pct=round(n_crash / N * 100, 1),
        collision_free_pct=round((N - n_crash) / N * 100, 1),
        avg_bat_used=round(float(d["bat_u"].mean()), 1)
            if "bat_u" in d.columns else 0.0,
        avg_efficiency=round(float(d["eff"].median()), 3)
            if "eff" in d.columns else 0.0,
    )


# ── Chart data builders ───────────────────────────────────────────────────────
def _s(v):
    if isinstance(v, float) and (np.isnan(v) or np.isinf(v)): return None
    if isinstance(v, (np.integer,)): return int(v)
    if isinstance(v, (np.floating,)): return float(v)
    return v

def outcome_bars(d):
    sc = d["flight_status"].value_counts().reset_index()
    sc.columns = ["status", "count"]
    sc = sc.sort_values("count", ascending=True)
    N  = len(d)
    return {"statuses": sc["status"].tolist(),
            "counts":   sc["count"].tolist(),
            "pcts":     [round(v / N * 100, 1) for v in sc["count"]]}

def coll_timeline(c):
    if c.empty or "gtn" not in c.columns: return {"events": []}
    events = []
    for _, row in c.iterrows():
        events.append({
            "tick":  _s(row["gtn"]),
            "type":  str(row.get("type", "—")),
            "severity": str(row.get("severity", "—")),
            "drone_a":  _s(row.get("drone_a_id")),
            "drone_b":  _s(row.get("drone_b_id")),
            "veh_a":    str(row.get("drone_a_vehicle", "—")),
            "veh_b":    str(row.get("drone_b_vehicle", "—")),
            "x":     _s(row.get("drone_a_coord_x")),
            "y":     _s(row.get("drone_a_coord_y")),
            "layer": _s(row.get("drone_a_layer")),
        })
    return {"events": events}

def drone_positions(d, c):
    pos    = d.dropna(subset=["cx", "cy"])
    drones = [{"id": _s(r["drone_id"]), "x": _s(r["cx"]), "y": _s(r["cy"]),
               "layer": _s(r.get("layer")), "alt": _s(r.get("alt")),
               "status": str(r["flight_status"]),
               "vehicle": str(r.get("vehicle", "—")),
               "bat_u": _s(r.get("bat_u")), "eff": _s(r.get("eff"))}
              for _, r in pos.iterrows()]
    conflicts = []
    if not c.empty:
        for _, r in c.iterrows():
            conflicts.append({
                "x":      _s(r.get("drone_a_coord_x")),
                "y":      _s(r.get("drone_a_coord_y")),
                "bx":     _s(r.get("drone_b_coord_x")),
                "by":     _s(r.get("drone_b_coord_y")),
                "type":   str(r.get("type", "—")),
                "tick":   _s(r.get("gtn")),
                "drone_a": _s(r.get("drone_a_id")),
                "drone_b": _s(r.get("drone_b_id")),
                "veh_a":  str(r.get("drone_a_vehicle", "—")),
                "veh_b":  str(r.get("drone_b_vehicle", "—")),
                "layer":  _s(r.get("drone_a_layer")),
            })
    return {"drones": drones, "conflicts": conflicts}

def battery_kde(d):
    from scipy.stats import gaussian_kde
    result = {}
    for field, key in [("bat_s", "start"), ("bat_u", "consumed"), ("bat_e", "end")]:
        vals = d[field].dropna().values if field in d.columns else np.array([])
        if len(vals) < 3:
            result[key] = {"x": [], "y": []}
            continue
        try:
            kde = gaussian_kde(vals, bw_method=0.3)
            xs  = np.linspace(max(0, vals.min()-2), min(100, vals.max()+2), 200)
            result[key] = {"x": xs.tolist(), "y": kde(xs).tolist()}
        except:
            result[key] = {"x": [], "y": []}
    return result


def battery_kde_by_status(d):
    """Per-flight-status battery consumed KDE curves for fleet page."""
    from scipy.stats import gaussian_kde
    result = []
    if "bat_u" not in d.columns or "flight_status" not in d.columns:
        return result
    for status, grp in d.groupby("flight_status"):
        vals = grp["bat_u"].dropna().values
        if len(vals) < 2:
            continue
        try:
            kde = gaussian_kde(vals, bw_method=0.4)
            xs  = np.linspace(max(0, vals.min()-2), vals.max()+2, 200)
            result.append({"status": status, "x": xs.tolist(), "y": kde(xs).tolist()})
        except:
            pass
    return result

def eff_scatter(d):
    ed = d.dropna(subset=["dp", "da"])
    ed = ed[ed["da"] > 0]
    if ed.empty: return {"points": [], "trend_x": [], "trend_y": []}
    pts = [{"dp": _s(r["dp"]), "da": _s(r["da"]),
             "status": str(r["flight_status"]),
             "vehicle": str(r.get("vehicle", "—")),
             "eff": _s(r.get("eff"))} for _, r in ed.iterrows()]
    try:
        z  = np.polyfit(ed["dp"].values, ed["da"].values, 1)
        tx = np.linspace(0, ed["dp"].max(), 60)
        return {"points": pts, "trend_x": tx.tolist(),
                "trend_y": np.polyval(z, tx).tolist()}
    except:
        return {"points": pts, "trend_x": [], "trend_y": []}

def vehicle_matrix(d):
    if "vehicle" not in d.columns:
        return {"vehicles": [], "bins": [], "z": []}
    tmp = d.dropna(subset=["spd", "vehicle"]).copy()
    tmp["spd_bin"] = pd.cut(tmp["spd"],
        bins=[0, 0.5, 1.0, 1.5, 2.0, 5.0],
        labels=["<0.5", "0.5-1", "1-1.5", "1.5-2", ">2"])
    if tmp["spd_bin"].isna().all():
        return {"vehicles": [], "bins": [], "z": []}
    mx = pd.crosstab(tmp["vehicle"], tmp["spd_bin"])
    return {"vehicles": list(mx.index),
            "bins": [str(c) for c in mx.columns],
            "z": mx.values.tolist()}

def rolling_rate(c):
    if c.empty or "gtn" not in c.columns:
        return {"ticks": [], "raw": [], "rolling": []}
    ct = c.dropna(subset=["gtn"])
    if ct.empty: return {"ticks": [], "raw": [], "rolling": []}
    gmin, gmax = int(ct["gtn"].min()), int(ct["gtn"].max())
    ticks = list(range(gmin, gmax + 1))
    tc    = ct["gtn"].value_counts()
    raw   = [int(tc.get(t, 0)) for t in ticks]
    roll  = pd.Series(raw).rolling(5, min_periods=1).mean().round(2).tolist()
    return {"ticks": ticks, "raw": raw, "rolling": roll}

def multitrail_data(DF_D, DF_C):
    rows = []
    for tid in DF_D["trial_id"].unique().tolist():
        td  = DF_D[DF_D["trial_id"] == tid]
        tc  = DF_C[DF_C["trial_id"] == tid] if not DF_C.empty else pd.DataFrame()
        n_  = len(td)
        ok  = int(td["flight_status"].str.startswith("Complete").sum()) if n_ else 0
        bu  = float(td["battery_consumed"].apply(parse_battery).mean()) if n_ else 0.0
        dp  = float(pd.to_numeric(td["distance_planned"], errors="coerce").mean()) if n_ else 0.0
        da  = float(pd.to_numeric(td["distance_actual"],  errors="coerce").mean()) if n_ else 0.0
        rows.append({
            "trial":       tid,
            "fleet":       n_,
            "complete":    ok,
            "comp_pct":    round(ok / n_ * 100, 1) if n_ else 0,
            "collisions":  len(tc),
            "coll_rate":   round(len(tc) / n_ * 100, 2) if n_ else 0,
            "bat_used":    round(bu, 1),
            "efficiency":  round(dp / da, 3)
                           if da and not np.isnan(da) and da > 0 else 0,
        })
    return {"trials": rows}



def layer_vehicle_heatmap(c):
    """Layer x vehicle collision count matrix for safety page."""
    if c.empty: return {"layers":[], "vehicles":[], "z":[]}
    import pandas as _pd
    av = _pd.concat([
        c[["drone_a_layer","drone_a_vehicle"]].rename(
            columns={"drone_a_layer":"layer","drone_a_vehicle":"vehicle"}),
        c[["drone_b_layer","drone_b_vehicle"]].rename(
            columns={"drone_b_layer":"layer","drone_b_vehicle":"vehicle"}),
    ]).reset_index(drop=True).dropna()
    if av.empty: return {"layers":[], "vehicles":[], "z":[]}
    grp   = av.groupby(["layer","vehicle"]).size().reset_index(name="count")
    layers= sorted(grp["layer"].unique().tolist())
    vehs  = sorted(grp["vehicle"].unique().tolist())
    z     = [[int(grp[(grp["layer"]==l)&(grp["vehicle"]==v)]["count"].sum())
              for v in vehs] for l in layers]
    return {"layers":[int(l) for l in layers], "vehicles": vehs, "z": z}


def cascade_gap_data(c):
    """Tick gaps between consecutive collision events for cascade analysis."""
    if c.empty or "gtn" not in c.columns: return []
    sorted_c = c.sort_values("gtn").dropna(subset=["gtn"])
    result = []
    for i, (_, row) in enumerate(sorted_c.iterrows()):
        result.append({
            "tick":     _s(row["gtn"]),
            "type":     str(row.get("type","—")),
            "severity": str(row.get("severity","—")),
            "drone_a":  _s(row.get("drone_a_id")),
            "drone_b":  _s(row.get("drone_b_id")),
            "veh_a":    str(row.get("drone_a_vehicle","—")),
            "veh_b":    str(row.get("drone_b_vehicle","—")),
        })
    return result


def bat_drain_points(d):
    """Per-drone battery consumed vs distance for fleet drain chart."""
    bd = d.dropna(subset=["da","bat_u","vehicle"])
    bd = bd[bd["da"] > 0]
    return [{"da": _s(r["da"]), "bat_u": _s(r["bat_u"]),
             "vehicle": str(r["vehicle"]),
             "status": str(r["flight_status"])} for _, r in bd.iterrows()]


def rerouting_overhead_points(d):
    """Efficiency per drone for rerouting overhead histogram."""
    ed = d.dropna(subset=["dp","da","eff"])
    ed = ed[(ed["da"]>0) & np.isfinite(ed["eff"])]
    return [{"eff": _s(r["eff"]), "status": str(r["flight_status"]),
             "dp": _s(r["dp"]), "da": _s(r["da"])} for _, r in ed.iterrows()]


def bat_reserve_layer_points(d):
    """Battery end per drone grouped by layer for box plot."""
    bd = d.dropna(subset=["layer","bat_e"])
    return [{"layer": _s(r["layer"]), "bat_e": _s(r["bat_e"]),
             "flight_status": str(r["flight_status"])} for _, r in bd.iterrows()]


def fleet_density_data(d, c):
    """Airborne drone count per grid tick vs collision count per tick."""
    if c.empty or "gtn" not in c.columns:
        return {"ticks":[], "airborne":[], "collision_counts":[]}
    c_t = c.dropna(subset=["gtn"])
    if c_t.empty: return {"ticks":[], "airborne":[], "collision_counts":[]}
    gmin, gmax = int(c_t["gtn"].min()), int(c_t["gtn"].max())
    ticks = list(range(gmin, gmax+1))
    # Estimate airborne: drones that haven't yet crashed or completed by this tick
    # Use grid_tick as a proxy — drones still "active" are those with higher ticks
    # Simple approach: total drones minus those that ended by each tick
    total = len(d)
    coll_counts = c_t["gtn"].value_counts()
    # Airborne approximation: decreases as collisions accumulate
    cumulative_lost = 0
    airborne = []
    coll_list = []
    for t in ticks:
        evts = int(coll_counts.get(t, 0))
        coll_list.append(evts)
        cumulative_lost += evts * 2  # each event removes 2 drones
        airborne.append(max(0, total - cumulative_lost))
    return {"ticks": ticks, "airborne": airborne, "collision_counts": coll_list}


def path_run_consistency(d, c):
    """Per-path-run KPIs for consistency analysis."""
    if "path_run" not in d.columns: return []
    rows = []
    for pr, grp in d.groupby("path_run"):
        n_ = len(grp)
        ok = int(grp["flight_status"].str.startswith("Complete").sum())
        tc = c[c["path_run"]==pr] if not c.empty and "path_run" in c.columns else pd.DataFrame()
        rows.append({
            "path_run":   int(pr),
            "comp_pct":   round(ok/n_*100,1) if n_ else 0,
            "coll_count": len(tc),
            "bat_avg":    round(float(grp["bat_u"].mean()),1) if "bat_u" in grp.columns else 0,
            "eff_median": round(float(grp["eff"].median()),3) if "eff" in grp.columns else 0,
        })
    return sorted(rows, key=lambda x: x["path_run"])


def bat_by_layer_data(d):
    """Battery consumed values per layer for violin/box chart."""
    result = {}
    if "layer" not in d.columns or "bat_u" not in d.columns: return result
    for ln in [1,2,3,4]:
        vals = d[d["layer"]==ln]["bat_u"].dropna().values
        result[str(int(ln))] = [_s(v) for v in vals]
    return result


def ml_risk_analysis(d):
    """Random Forest risk scoring + anomaly detection per drone."""
    try:
        from sklearn.ensemble import RandomForestClassifier, IsolationForest
        import warnings; warnings.filterwarnings("ignore")
    except ImportError:
        return {"error": "scikit-learn not installed. Run: pip install scikit-learn",
                "drones":[], "feature_importance":{}, "risk_tiers":{}}

    features = ["bat_s","cx","cy","layer","dp"]
    df = d.dropna(subset=features+["flight_status"]).copy()
    if len(df) < 10:
        return {"error":"Insufficient data","drones":[],"feature_importance":{},"risk_tiers":{}}

    df["crashed"] = df["flight_status"].str.startswith("Collision").astype(int)
    X = df[features].values
    y = df["crashed"].values
    feat_labels = ["Battery Start","X Coord","Y Coord","Layer","Planned Distance"]

    # Risk scoring
    rf = RandomForestClassifier(n_estimators=100, random_state=42,
                                class_weight="balanced")
    rf.fit(X, y)
    probs = rf.predict_proba(X)[:,1]
    df["risk_score"] = probs

    # Anomaly detection
    iso = IsolationForest(contamination=0.15, random_state=42)
    df["anomaly"] = (iso.fit_predict(X) == -1).astype(int)

    # Risk tier
    df["risk_tier"] = "Low"
    df.loc[probs >= 0.3, "risk_tier"] = "Medium"
    df.loc[probs >= 0.6, "risk_tier"] = "High"

    # Build output
    drones_out = []
    for _, row in df.iterrows():
        drones_out.append({
            "id":         _s(row["drone_id"]),
            "x":          _s(row["cx"]),
            "y":          _s(row["cy"]),
            "layer":      _s(row["layer"]),
            "risk_score": round(float(row["risk_score"]),3),
            "risk_tier":  row["risk_tier"],
            "anomaly":    int(row["anomaly"]),
            "status":     str(row["flight_status"]),
            "bat_s":      _s(row["bat_s"]),
            "dp":         _s(row["dp"]),
            "crashed":    int(row["crashed"]),
        })

    fi = {feat_labels[i]: round(float(rf.feature_importances_[i]),3)
          for i in range(len(feat_labels))}
    tiers = df["risk_tier"].value_counts().to_dict()
    tier_counts = {"Low":int(tiers.get("Low",0)),
                   "Medium":int(tiers.get("Medium",0)),
                   "High":int(tiers.get("High",0))}

    # Risk score vs actual outcome confusion
    tp = int(((probs >= 0.5) & (y == 1)).sum())
    fp = int(((probs >= 0.5) & (y == 0)).sum())
    tn = int(((probs <  0.5) & (y == 0)).sum())
    fn = int(((probs <  0.5) & (y == 1)).sum())

    return {
        "drones":            drones_out,
        "feature_importance": fi,
        "risk_tiers":        tier_counts,
        "total_analyzed":    len(df),
        "confusion":         {"tp":tp,"fp":fp,"tn":tn,"fn":fn},
        "n_anomalies":       int(df["anomaly"].sum()),
    }


def full_event_timeline(d, c):
    """All drone events across all statuses for complete timeline."""
    events = []
    # Drone-level terminal events from drone summary
    status_lane = {
        "Complete":                  0,
        "Incomplete — Battery":      1,
        "Cancelled — Pre-flight":    2,
        "Cancelled — In-flight":     3,
    }
    for _, row in d.iterrows():
        st = str(row.get("flight_status",""))
        if st in status_lane:
            events.append({
                "lane":     status_lane[st],
                "status":   st,
                "id":       _s(row.get("drone_id")),
                "duration": _s(row.get("duration")),
                "bat_u":    _s(row.get("bat_u")),
                "vehicle":  str(row.get("vehicle","—")),
                "layer":    _s(row.get("layer")),
            })

    # Collision events with full severity breakdown
    coll_lane = {
        "Direct":    4,
        "Proximity": 5,
        "Near Miss": 6,
    }
    sev_lane = {
        "Critical":  4,
        "Major":     4,
        "Minor":     5,
        "Near Miss": 6,
    }
    if not c.empty:
        for _, row in c.iterrows():
            ctype = str(row.get("type","—"))
            sev   = str(row.get("severity","—"))
            events.append({
                "lane":     sev_lane.get(sev, coll_lane.get(ctype, 6)),
                "status":   sev,
                "tick":     _s(row.get("gtn")),
                "id_a":     _s(row.get("drone_a_id")),
                "id_b":     _s(row.get("drone_b_id")),
                "veh_a":    str(row.get("drone_a_vehicle","—")),
                "type":     ctype,
                "severity": sev,
                "x":        _s(row.get("drone_a_coord_x")),
                "y":        _s(row.get("drone_a_coord_y")),
                "layer":    _s(row.get("drone_a_layer")),
            })
    return {"events": events}


def tick_event_heatmap(c):
    """Tick x event-type count grid for event intensity heatmap."""
    if c.empty or "gtn" not in c.columns:
        return {"ticks":[], "types":[], "z":[]}
    c_t = c.dropna(subset=["gtn"])
    if c_t.empty:
        return {"ticks":[], "types":[], "z":[]}
    ticks   = sorted(c_t["gtn"].unique().tolist())
    types   = ["Critical","Minor","Near Miss"]
    sev_map = c_t.groupby(["gtn","severity"]).size().reset_index(name="count")
    z = []
    for sev in types:
        row_data = []
        for t in ticks:
            val = sev_map[(sev_map["gtn"]==t)&(sev_map["severity"]==sev)]["count"]
            row_data.append(int(val.sum()) if len(val) else 0)
        z.append(row_data)
    return {"ticks":[int(t) for t in ticks], "types": types, "z": z}


def fleet_outcome_funnel(d):
    """Waterfall breakdown of fleet from launch to terminal states."""
    N = len(d)
    complete   = int(d["flight_status"].str.startswith("Complete").sum())
    coll_direct= int((d["flight_status"]=="Collision — Node to Node").sum())
    coll_prox  = int((d["flight_status"]=="Collision — Proximity").sum())
    batt_fail  = int((d["flight_status"]=="Incomplete — Battery").sum())
    canc_inf   = int((d["flight_status"]=="Cancelled — In-flight").sum())
    canc_pre   = int((d["flight_status"]=="Cancelled — Pre-flight").sum())
    failed = N - complete
    return {
        "total":        N,
        "complete":     complete,
        "coll_direct":  coll_direct,
        "coll_prox":    coll_prox,
        "batt_fail":    batt_fail,
        "canc_inflight":canc_inf,
        "canc_preflight":canc_pre,
        "failed":       failed,
        "success_rate": round(complete/N*100,1) if N else 0,
    }


def cumulative_severity_progression(c):
    """Cumulative collision events by severity over grid ticks."""
    if c.empty or "gtn" not in c.columns:
        return {"ticks":[], "critical":[], "minor":[], "near_miss":[]}
    c_t = c.dropna(subset=["gtn"])
    if c_t.empty:
        return {"ticks":[], "critical":[], "minor":[], "near_miss":[]}
    gmin, gmax = int(c_t["gtn"].min()), int(c_t["gtn"].max())
    ticks = list(range(1, gmax+1))
    sev_counts = c_t.groupby(["gtn","severity"]).size().reset_index(name="count")
    cum = {"Critical":0,"Minor":0,"Near Miss":0}
    crit_list, minor_list, nm_list = [], [], []
    for t in ticks:
        for sev in ["Critical","Minor","Near Miss"]:
            val = sev_counts[(sev_counts["gtn"]==t)&(sev_counts["severity"]==sev)]["count"]
            cum[sev] += int(val.sum()) if len(val) else 0
        crit_list.append(cum["Critical"])
        minor_list.append(cum["Minor"])
        nm_list.append(cum["Near Miss"])
    return {"ticks":ticks,"critical":crit_list,"minor":minor_list,"near_miss":nm_list}


def zone_crash_heatmap(d):
    """5x5 zone grid crash rate heatmap for mission success spatial analysis."""
    import pandas as _pd
    d2 = d.dropna(subset=["cx","cy"]).copy()
    if d2.empty: return {"z":[],"text":[],"x_labels":[],"y_labels":[]}
    d2["crashed"] = d2["flight_status"].str.startswith("Collision").astype(int)
    d2["zx"] = (d2["cx"]//10).clip(0,4).astype(int)
    d2["zy"] = (d2["cy"]//10).clip(0,4).astype(int)
    zones = d2.groupby(["zx","zy"]).agg(
        crash_rate=("crashed","mean"),
        count=("crashed","count"),
        n_crashes=("crashed","sum")
    ).reset_index()
    z    = [[None]*5 for _ in range(5)]
    text = [["—"]*5 for _ in range(5)]
    for _, row in zones.iterrows():
        xi, yi = int(row["zx"]), int(row["zy"])
        if 0<=xi<5 and 0<=yi<5:
            z[yi][xi]    = round(float(row["crash_rate"]),2)
            text[yi][xi] = f"{int(row['n_crashes'])}/{int(row['count'])}"
    labels = ["0-10","10-20","20-30","30-40","40-50"]
    return {"z":z,"text":text,"x_labels":labels,"y_labels":labels}


def collision_pair_network(d, c):
    """Node-link network of collision pairs."""
    if c.empty: return {"nodes":[],"edges":[]}
    pos = {}
    for _, row in d.dropna(subset=["cx","cy"]).iterrows():
        pos[row["drone_id"]] = {
            "x":float(row["cx"]),"y":float(row["cy"]),
            "status":str(row["flight_status"]),
            "vehicle":str(row.get("vehicle","—")),
        }
    crashed_ids = set(d[d["flight_status"].str.startswith("Collision")]["drone_id"].tolist())
    edges = []
    involved = set()
    for _, row in c.iterrows():
        a, b = row["drone_a_id"], row["drone_b_id"]
        involved.add(a); involved.add(b)
        edges.append({"a":int(a),"b":int(b),"type":str(row.get("type","—")),
                      "tick":_s(row.get("gtn"))})
    nodes = []
    for did in involved:
        p = pos.get(did,{})
        deg = sum(1 for e in edges if e["a"]==did or e["b"]==did)
        nodes.append({"id":int(did),
                      "x":p.get("x",25.0),"y":p.get("y",25.0),
                      "crashed":did in crashed_ids,
                      "vehicle":p.get("vehicle","—"),
                      "status":p.get("status","—"),
                      "degree":deg})
    return {"nodes":nodes,"edges":edges}


def layer_collision_density(d, c):
    """Collision count normalized by drone count per layer."""
    import pandas as _pd
    layer_counts = d["layer"].value_counts().to_dict()
    coll_by_layer = (_pd.concat([c["drone_a_layer"],c["drone_b_layer"]])
                     .value_counts().to_dict()) if not c.empty else {}
    rows = []
    for l in [1,2,3,4]:
        n_d = int(layer_counts.get(float(l), layer_counts.get(l, 0)))
        n_c = int(coll_by_layer.get(float(l), coll_by_layer.get(l, 0)))
        rows.append({"layer":l,"label":f"L{l} · {(l-1)*50}m",
                     "drones":n_d,"collisions":n_c,
                     "density":round(n_c/n_d,3) if n_d>0 else 0})
    return {"layers":rows}


def vehicle_radar_data(d, c=None):
    """5-metric dot matrix using highest-variance metrics across vehicle types.
    Axes: Completion, Battery Fail Rate (inverted), Avg Battery Consumed (inverted),
          Avg Battery Remaining, Cancellation Rate (inverted)
    All normalised 0-100 so bigger dot = better performance.
    """
    rows = []
    for veh, grp in d.groupby("vehicle"):
        n = len(grp)
        if n == 0: continue

        # 1. Completion rate — direct
        comp = round(grp["flight_status"].str.startswith("Complete").sum()/n*100, 1)

        # 2. Battery failure rate — inverted (lower fail = better)
        bat_fail = round((grp["flight_status"]=="Incomplete — Battery").sum()/n*100, 1)
        bat_fail_score = round(100 - bat_fail, 1)

        # 3. Avg battery consumed — inverted (lower consumption = better)
        bat_u = round(float(grp["bat_u"].mean()), 1)
        bat_u_score = round(100 - bat_u, 1)

        # 4. Avg battery remaining at end — direct (more reserve = better)
        bat_e = round(float(grp["bat_e"].mean()), 1)

        # 5. Cancellation rate — inverted (fewer cancellations = better)
        canc = round(grp["flight_status"].str.startswith("Cancelled").sum()/n*100, 1)
        canc_score = round(100 - canc, 1)

        rows.append({
            "vehicle":     veh,
            "completion":  comp,
            "bat_fail":    bat_fail_score,
            "bat_consumed": bat_u_score,
            "bat_reserve": bat_e,
            "cancellation": canc_score,
            "raw": {
                "comp":     comp,
                "bat_fail": bat_fail,
                "bat_u":    bat_u,
                "bat_e":    bat_e,
                "canc":     canc,
            },
        })
    return {
        "vehicles": rows,
        "axes": [
            "Mission Completion",
            "Battery Endurance",
            "Energy Efficiency",
            "Reserve at Landing",
            "Launch Reliability",
        ],
    }


def route_length_by_vehicle(d):
    """Route length bucket distribution per vehicle type."""
    import pandas as _pd
    d2 = d.dropna(subset=["dp","vehicle"]).copy()
    if d2.empty: return {"vehicles":[],"buckets":[],"data":{}}
    d2["bucket"] = _pd.cut(d2["dp"],
        bins=[0,15,30,45,200],
        labels=["Short (0-15)","Medium (15-30)","Long (30-45)","Very Long (45+)"])
    ct = _pd.crosstab(d2["vehicle"], d2["bucket"])
    return {
        "vehicles": list(ct.index),
        "buckets":  [str(c) for c in ct.columns],
        "data":     {veh: ct.loc[veh].tolist() for veh in ct.index},
    }


def layer_safety_profile(d):
    """Completion rate, crash rate, battery per altitude layer."""
    rows = []
    for l in [1,2,3,4]:
        grp = d[d["layer"]==float(l)]
        if len(grp) == 0: continue
        n     = len(grp)
        comp  = round(grp["flight_status"].str.startswith("Complete").sum()/n*100, 1)
        crash = round(grp["flight_status"].str.startswith("Collision").sum()/n*100, 1)
        batt  = round(grp["flight_status"].str.startswith("Incomplete").sum()/n*100, 1)
        canc  = round(grp["flight_status"].str.startswith("Cancelled").sum()/n*100, 1)
        rows.append({
            "layer":   l,
            "label":   f"L{l} · {(l-1)*50}m",
            "drones":  n,
            "complete_pct": comp,
            "crash_pct":    crash,
            "battery_pct":  batt,
            "cancel_pct":   canc,
        })
    return {"layers": rows}


def conflict_escalation(c):
    """
    Track drone pairs that escalated from Near Miss / Minor → Direct crash.
    Shows which conflicts the algorithm tried to resolve but failed.
    Returns timeline of events per drone pair, sorted by first encounter tick.
    """
    if c.empty: return {"pairs": [], "ticks": []}
    import pandas as _pd

    # Build pair key (always smaller ID first for consistency)
    c2 = c.copy()
    c2["pair"] = c2.apply(
        lambda r: f"D{min(int(r.drone_a_id),int(r.drone_b_id))}-D{max(int(r.drone_a_id),int(r.drone_b_id))}",
        axis=1
    )

    pairs_out = []
    for pair, grp in c2.groupby("pair"):
        grp = grp.sort_values("gtn")
        events = []
        for _, row in grp.iterrows():
            events.append({
                "tick":     int(row["gtn"]),
                "type":     str(row["type"]),
                "severity": str(row["severity"]),
            })

        # Classify escalation
        sevs = [e["severity"] for e in events]
        escalated = ("Near Miss" in sevs or "Minor" in sevs) and "Critical" in sevs
        final = sevs[-1]

        pairs_out.append({
            "pair":      pair,
            "events":    events,
            "n_events":  len(events),
            "escalated": escalated,
            "final":     final,
            "first_tick": events[0]["tick"],
            "last_tick":  events[-1]["tick"],
            "span":       events[-1]["tick"] - events[0]["tick"],
        })

    # Sort: escalated pairs first, then by first tick
    pairs_out.sort(key=lambda x: (not x["escalated"], x["first_tick"]))

    # All unique ticks for X axis
    all_ticks = sorted(c2["gtn"].unique().tolist())
    return {"pairs": pairs_out, "ticks": [int(t) for t in all_ticks]}

# ── Helpers ───────────────────────────────────────────────────────────────────
def _get(tid: str, path_runs: Optional[str] = None):
    if _S["drones"].empty:
        raise HTTPException(400, "No data loaded")
    d = _S["drones"][_S["drones"]["trial_id"] == tid].copy()
    c = (_S["collisions"][_S["collisions"]["trial_id"] == tid].copy()
         if not _S["collisions"].empty else pd.DataFrame())
    if path_runs:
        prs = [int(x) for x in path_runs.split(",") if x.strip().isdigit()]
        if prs:
            d = d[d["path_run"].isin(prs)]
            if not c.empty and "path_run" in c.columns:
                c = c[c["path_run"].isin(prs)]
    return enrich(d, c)


# ── Routes ────────────────────────────────────────────────────────────────────
@app.get("/")
async def root():
    return FileResponse(str(FRONTEND / "index.html"))

@app.post("/api/connect/postgres")
async def connect_pg(
    host: str = Form("localhost"), port: int = Form(5432),
    dbname: str = Form("simulation_db"), user: str = Form("postgres"),
    password: str = Form("omkar7781"),
):
    try:
        data = load_postgres(host, port, dbname, user, password)
        _S["drones"]     = data["drones"]
        _S["collisions"] = data["collisions"]
        _S["source"]     = "postgresql"
        _S["trials"]     = sorted(_S["drones"]["trial_id"].unique().tolist())
        return {"ok": True, "trials": _S["trials"], "source": "postgresql"}
    except Exception as e:
        raise HTTPException(400, str(e))

@app.post("/api/connect/excel")
async def connect_excel(file: UploadFile = File(...)):
    try:
        buf = io.BytesIO(await file.read())
        buf.name = file.filename
        data = load_excel(buf)
        _S["drones"]     = data["drones"]
        _S["collisions"] = data["collisions"]
        _S["source"]     = data["source"]
        _S["trials"]     = sorted(_S["drones"]["trial_id"].unique().tolist())
        return {"ok": True, "trials": _S["trials"], "source": data["source"]}
    except Exception as e:
        raise HTTPException(400, str(e))

@app.get("/api/trials")
async def get_trials():
    return {"trials": _S["trials"], "source": _S["source"]}

@app.get("/api/trial/{tid}/path_runs")
async def get_path_runs(tid: str):
    if _S["drones"].empty: return {"path_runs": []}
    prs = sorted(_S["drones"][_S["drones"]["trial_id"] == tid]["path_run"]
                 .dropna().unique().tolist())
    return {"path_runs": [int(p) for p in prs]}

@app.get("/api/trial/{tid}/kpis")
async def get_kpis(tid: str, path_runs: Optional[str] = None):
    d, c = _get(tid, path_runs)
    return J(compute_kpis(d, c))

@app.get("/api/trial/{tid}/overview")
async def get_overview(tid: str, path_runs: Optional[str] = None):
    d, c = _get(tid, path_runs)
    k    = compute_kpis(d, c)
    layers   = {str(ln): int((d["layer"] == ln).sum())
                for ln in [1, 2, 3, 4]} if "layer" in d.columns else {}
    vehicles = {str(v): int(cnt)
                for v, cnt in d["vehicle"].value_counts().items()} \
               if "vehicle" in d.columns else {}
    coll_log = []
    if not c.empty:
        for _, row in c.head(30).iterrows():
            coll_log.append({
                "path_run": _s(row.get("path_run")),
                "tick":     _s(row.get("gtn")),
                "type":     str(row.get("type", "—")),
                "severity": str(row.get("severity", "—")),
                "drone_a":  _s(row.get("drone_a_id")),
                "drone_b":  _s(row.get("drone_b_id")),
                "veh_a":    str(row.get("drone_a_vehicle", "—")),
                "veh_b":    str(row.get("drone_b_vehicle", "—")),
                "x":        _s(row.get("drone_a_coord_x")),
                "y":        _s(row.get("drone_a_coord_y")),
                "layer":    _s(row.get("drone_a_layer")),
            })
    return J({"kpis": k, "outcomes": outcome_bars(d),
              "layers": layers, "vehicles": vehicles, "coll_log": coll_log})

@app.get("/api/trial/{tid}/spatial")
async def get_spatial(tid: str, path_runs: Optional[str] = None):
    d, c = _get(tid, path_runs)
    return J(drone_positions(d, c))

@app.get("/api/trial/{tid}/safety")
async def get_safety(tid: str, path_runs: Optional[str] = None):
    d, c = _get(tid, path_runs)
    if c.empty:
        return J({"events": [], "severity": [], "vehicle_hits": [],
                  "bat_kde": {"x": [], "y": [], "obs": []},
                  "layer_veh_heatmap": {"layers":[],"vehicles":[],"z":[]},
                  "cascade_events": [],
                  "pair_network": {"nodes":[],"edges":[]},
                  "layer_density": {"layers":[]}})
    by_run   = c.groupby(["path_run", "type"]).size().reset_index(name="count")
    sev      = c["severity"].value_counts().reset_index()
    sev.columns = ["severity", "count"]
    veh_hits = pd.concat([c["drone_a_vehicle"],
                          c["drone_b_vehicle"]]).value_counts().reset_index()
    veh_hits.columns = ["vehicle", "collisions"]
    fc = d["vehicle"].value_counts().to_dict() if "vehicle" in d.columns else {}
    veh_hits["fleet"]    = veh_hits["vehicle"].map(fc).fillna(0)
    veh_hits["rate_pct"] = (veh_hits["collisions"] /
                            veh_hits["fleet"].replace(0, np.nan) * 100).round(1)
    bat_crash = pd.concat([c["bat_a"].rename("b"),
                           c["bat_b"].rename("b")]).dropna().values
    bat_kde_out = {"x": [], "y": []}
    if len(bat_crash) >= 3:
        try:
            from scipy.stats import gaussian_kde
            kde = gaussian_kde(bat_crash, bw_method=0.4)
            xs  = np.linspace(0, 100, 200)
            bat_kde_out = {"x": xs.tolist(), "y": kde(xs).tolist(),
                           "obs": bat_crash.tolist()}
        except: pass
    dist_crash = []
    for _, row in c.iterrows():
        for f in ["da_a", "da_b"]:
            v = row.get(f)
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                dist_crash.append({"type": str(row.get("type", "—")),
                                   "dist": float(v)})
    return J({
        "events":            coll_timeline(c)["events"],
        "severity":          sev.to_dict("records"),
        "vehicle_hits":      veh_hits.to_dict("records"),
        "bat_kde":           bat_kde_out,
        "layer_veh_heatmap": layer_vehicle_heatmap(c),
        "cascade_events":    cascade_gap_data(c),
        "layer_safety":      layer_safety_profile(d),
        "route_by_vehicle":  route_length_by_vehicle(d),
    })

@app.get("/api/trial/{tid}/fleet")
async def get_fleet(tid: str, path_runs: Optional[str] = None):
    d, c = _get(tid, path_runs)
    k    = compute_kpis(d, c)
    veh_summary = []
    if "vehicle" in d.columns:
        for veh, grp in d.groupby("vehicle"):
            nv = len(grp)
            ok = int(grp["flight_status"].str.startswith("Complete").sum())
            cr = int(grp["flight_status"].str.startswith("Collision").sum())
            veh_summary.append({
                "vehicle":  veh, "fleet": nv, "complete": ok,
                "comp_pct": round(ok / nv * 100, 1),
                "collisions": cr, "crash_pct": round(cr / nv * 100, 1),
                "avg_bat":  round(float(grp["bat_u"].mean()), 1),
                "avg_eff":  round(float(grp["eff"].median()), 3),
            })
    dur  = d.dropna(subset=["duration"]) if "duration" in d.columns else pd.DataFrame()
    dur  = dur[dur["duration"] > 0] if not dur.empty else dur
    dur_by_status = [{"status": s, "values": grp["duration"].tolist()}
                     for s, grp in dur.groupby("flight_status")]                     if not dur.empty else []
    return J({
        "kpis":               k,
        "bat_kde":            battery_kde(d),
        "bat_kde_by_status":  battery_kde_by_status(d),
        "eff_scatter":        eff_scatter(d),
        "bat_drain_points":   bat_drain_points(d),
        "bat_reserve_layer":  bat_reserve_layer_points(d),
        "fleet_funnel":       fleet_outcome_funnel(d),
        "dur_by_status":      dur_by_status,
        "veh_summary":        veh_summary,
        "zone_crash":         zone_crash_heatmap(d),
        "vehicle_radar":      vehicle_radar_data(d),
        "route_by_vehicle":   route_length_by_vehicle(d),
    })

@app.get("/api/trial/{tid}/temporal")
async def get_temporal(tid: str, path_runs: Optional[str] = None):
    d, c = _get(tid, path_runs)
    dur  = d.dropna(subset=["duration"]) if "duration" in d.columns else pd.DataFrame()
    dur  = dur[dur["duration"] > 0] if not dur.empty else dur
    dur_by_status = [{"status": s, "values": grp["duration"].tolist()}
                     for s, grp in dur.groupby("flight_status")] \
                    if not dur.empty else []
    return J({
        "full_timeline":  full_event_timeline(d, c),
        "fleet_density":  fleet_density_data(d, c),
        "dur_by_status":  dur_by_status,
    })

@app.get("/api/trial/{tid}/ml_risk")
async def get_ml_risk(tid: str, path_runs: Optional[str] = None):
    d, c = _get(tid, path_runs)
    return J({
        "ml":                ml_risk_analysis(d),
        "fleet_funnel":      fleet_outcome_funnel(d),
        "event_heatmap":     tick_event_heatmap(c),
        "cum_severity":      cumulative_severity_progression(c),
        "vehicle_radar":     vehicle_radar_data(d),
        "conflict_escalation": conflict_escalation(c),
    })


@app.get("/api/multitrail_intel")
async def get_multitrail_intel():
    DF_D = _S["drones"]; DF_C = _S["collisions"]
    if DF_D.empty: raise HTTPException(400, "No data loaded")
    trials = DF_D["trial_id"].unique().tolist()
    rows = []
    for tid in trials:
        td = DF_D[DF_D["trial_id"]==tid]
        tc = DF_C[DF_C["trial_id"]==tid] if not DF_C.empty else pd.DataFrame()
        n_ = len(td)
        ok = int(td["flight_status"].str.startswith("Complete").sum()) if n_ else 0
        bu = float(td["battery_consumed"].apply(parse_battery).mean()) if n_ else 0.0
        dp = float(pd.to_numeric(td["distance_planned"],errors="coerce").mean()) if n_ else 0.0
        da = float(pd.to_numeric(td["distance_actual"], errors="coerce").mean()) if n_ else 0.0
        d_enr, c_enr = enrich(td, tc)
        ml_res = ml_risk_analysis(d_enr)
        high_risk = ml_res.get("risk_tiers",{}).get("High",0)
        rows.append({
            "trial":            tid, "fleet": n_, "complete": ok,
            "comp_pct":         round(ok/n_*100,1) if n_ else 0,
            "collisions":       len(tc),
            "coll_rate":        round(len(tc)/n_*100,2) if n_ else 0,
            "bat_used":         round(bu,1),
            "efficiency":       round(dp/da,3) if da and not np.isnan(da) and da>0 else 0,
            "high_risk_drones": high_risk,
        })
    return J({"trials": rows, "n_trials": len(trials)})




if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
